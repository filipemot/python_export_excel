from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.formula.translate import Translator
from models.type_column_enum import TypeColumn

def get_active_sheet(active_sheet, sheet, file_workbook):
    """
    Get active sheet
    :param Object | active_sheet:
        The object with the instance of Active Sheet
    :param Object{SheetExcel} | sheet:
        The object of type SheetExcel with information of sheet
    :param Object{Workbook} | file_workbook:
        The object of type Workbook with information of woorkbook
    """
    try:
        if(active_sheet is None):
            active_sheet = file_workbook.active
            active_sheet.title = sheet.title
        else:
            active_sheet = file_workbook.create_sheet(sheet.title)

        return active_sheet
    except:
        print('get_active_sheet() - workbook_excel_service', 'Error')

def create_cell(active_sheet, cell, id_row, id_column):
    """
    Create a cell of sheet
    :param Object | active_sheet:
        The object with the instance of Active Sheet
    :param Object{CellSheetExcel} | cell:
        The object of type CellSheetExcel with information of cell
    :param int | id_row:
        The index of row
    :param int | id_column:
        The index of column
    """
    try:
        active_cell = active_sheet.cell(row=id_row, column=id_column)
        create_value(active_cell, cell)
    except:
        print('create_cell() - workbook_excel_service', 'Error')

def create_value(active_cell, cell):
    """
    Create value of cell
    :param Object | active_sheet:
        The object with the instance of Active Sheet
    :param Object{CellSheetExcel} | cell:
        The object of type CellSheetExcel with information of cell
    """
    try:
        active_cell.font = (Font(bold=cell.font_bold, size=cell.font_size))

        if(cell.type_column == TypeColumn.DATE or cell.type_column == TypeColumn.DECIMAL):
            active_cell.number_format = cell.format_value

        active_cell.value = cell.value
    except:
        print('create_value() - workbook_excel_service', 'Error')

def create_row(row, active_sheet, id_row):
    """
    Create a row of sheet
    :param Object{RowSheetExcel} | row:
        The object of type RowSheetExcel with information of row
    :param Object | active_sheet:
        The object with the instance of Active Sheet
    :param int | id_row:
        The index of row
    """
    try:
        id_column = 1
        for cell in row.cell_sheets:
            create_cell(active_sheet, cell, id_row, id_column)
            id_column = id_column + 1
    except:
        print('create_row() - workbook_excel_service', 'Error')

def create_sheet(file_workbook, active_sheet, sheet):
    """
    Create a sheet
    :param Object{Workbook} | file_workbook:
        The object of type Workbook with information of woorkbook
    :param Object | active_sheet:
        The object with the instance of Active Sheet
    :param Object{SheetExcel} | sheet:
        The object of type SheetExcel with information of sheet
    """
    try:
        active_sheet = get_active_sheet(active_sheet, sheet, file_workbook)
        id_row = 1

        for row in sheet.row_sheets:
            create_row(row, active_sheet, id_row)
            id_row = id_row + 1

        return active_sheet
    except:
        print('create_sheet() - workbook_excel_service', 'Error')

def create_workbook(workbook):
    """
    Create a workbook
    :param Object{WorkbookExcel} | workbook:
        The object of type WorkbookExcel with information of woorkbook
    """
    try:
        file_workbook = Workbook()
        active_sheet = None

        for sheet in workbook.sheets:
            active_sheet = create_sheet(file_workbook, active_sheet, sheet)

        file_workbook.save(workbook.filename)
    except:
        print('create_workbook() - workbook_excel_service', 'Error')